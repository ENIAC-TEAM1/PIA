import os
import re
import matplotlib.pyplot as plt
from collections import Counter
from datetime import datetime
import openpyxl

# Funcion para obtener el siguiente numero entre parentesis para manejar el nombre de los archivos(expresiones regulares)
def obtener_siguiente_numero_archivos(carpeta, patron):
    archivos_existentes = [f for f in os.listdir(carpeta) if os.path.isfile(os.path.join(carpeta, f)) and re.search(patron, f)]
    
    if archivos_existentes:
        numeros_en_archivos = [int(re.search(r'\((\d+)\)', archivo).group(1)) for archivo in archivos_existentes if re.search(r'\((\d+)\)', archivo)]
        siguiente_numero = max(numeros_en_archivos) + 1
    else:
        siguiente_numero = 1

    return siguiente_numero

# Funcion que retorna la cadena de cada region para las request del API
def obtener_cadena_por_region():
    while True:
        print("\033[1;94mSeleccione una region:\033[0m")
        print("\033[94ma)\033[0m Norteamerica")
        print("\033[94mb)\033[0m Brasil")
        print("\033[94mc)\033[0m Europa Oeste")
        print("\033[94md)\033[0m Latino America Norte")
        print("\033[94me)\033[0m Latino America Sur")

        opcion_region = input("\033[94mIngrese la opcion de la region: \033[0m").lower()

        if opcion_region in ['a', 'b', 'c', 'd', 'e']:
            break
        else:
            print("\033[91mOpcion no valida. Intente nuevamente.\033[0m")

    if opcion_region == 'a':
        return "na1"
    elif opcion_region == 'b':
        return "br1"
    elif opcion_region == 'c':
        return "EUW1"
    elif opcion_region == 'd':
        return "LA1"
    elif opcion_region == 'e':
        return "LA2"

# Funcion que grafica los datos de la funcion KDA
def graficar_estadisticas_partidas(nombre, menu_grafica, partidas, kills_lista, asistencias_lista, muertes_lista):
    #Datos de la grafica
    plt.plot(partidas, kills_lista, label='Kills', marker='o')
    plt.plot(partidas, asistencias_lista, label='Asistencias', marker='o')
    plt.plot(partidas, muertes_lista, label='Muertes', marker='o')

    #Informacion en la grafica
    plt.xlabel('Partidas')
    plt.ylabel('Cantidad')
    plt.title('Kills, Asistencias y Muertes en las ultimas partidas')
    plt.legend()


    carpeta_graficas = 'Graficas'
    prefijo_archivo = f"grafico-kda-{nombre}"
    patron_archivo = f"{prefijo_archivo}.*"
    try:  
        if menu_grafica == 'a' or menu_grafica == 'b':
            os.makedirs(carpeta_graficas, exist_ok=True)

            # Obtener el siguiente numero para el archivo
            siguiente_numero = obtener_siguiente_numero_archivos(carpeta_graficas, patron_archivo)
            
            fecha_hora_actual = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            
            # Guardar la figura en la carpeta
            ruta_archivo = os.path.join(carpeta_graficas, f"{prefijo_archivo}({siguiente_numero})_{fecha_hora_actual}.png")
            plt.savefig(ruta_archivo)
            print(F"\033[92mGRAFICA GUARDADA EXITOSAMENTE EN {ruta_archivo}\033[0m")
        if menu_grafica == 'a':
          plt.show()
        plt.close()

    except Exception as e:
        print(f"Error al guardar la grafica: {e}")

# Funcion que grafica los datos de la funcion Tasa de victorias y derrotas
def graficar_estadisticas_victorias(nombre, menu_grafica, sizes):
    # Datos de la grafica
    labels = 'Victorias', 'Derrotas'
    colors = ['green', 'red']
    explode = (0.1, 0)
    plt.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%', shadow=True, startangle=140)
    plt.axis('equal')
    plt.title('Porcentaje de Victorias y Derrotas')
    
    # Se crean cadenas por si se pide guardar la grafica
    carpeta_graficas = 'Graficas'
    prefijo_archivo = f"grafico-tasa-victoria-{nombre}"
    patron_archivo = f"{prefijo_archivo}.*"  
    try:
        # Si el usuario eligio guardar la funcion (a y b)
        if menu_grafica == 'a' or menu_grafica == 'b':
            
            # Se crea la carpeta si no existe
            os.makedirs(carpeta_graficas, exist_ok=True)
            
            # Obtener el siguiente numero para el nombre del archivo
            siguiente_numero = obtener_siguiente_numero_archivos(carpeta_graficas, patron_archivo)
            
            #   Obtiene la fecha y hora actual para el nombre del archivo
            fecha_hora_actual = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            
            # Guardar la figura en la carpeta
            ruta_archivo = os.path.join(carpeta_graficas, f"{prefijo_archivo}({siguiente_numero})_{fecha_hora_actual}.png")
            plt.savefig(ruta_archivo)
            print(F"\033[92mGRAFICA GUARDADA EXITOSAMENTE EN {ruta_archivo}\033[0m")
        if menu_grafica == 'a':
            plt.show()
        plt.close()
          
    except Exception as e:
        print(f"Error al guardar la grafica: {e}")

# Funcion que grafica los datos de la funcion Campeon mas usado
def graficar_campeones_usados(nombre, menu_grafica, campeones_usados):
    #Una vez que se reciben los campeones usados
    # Se cuenta la frecuencia de cada campeon con counter que regresa un diccionario
    conteo_campeones = Counter(campeones_usados)

    # Despues se separa el diccionario por keys y valkues para la grafica
    campeones = list(conteo_campeones.keys())
    frecuencias = list(conteo_campeones.values())

    # Luego se crea la grafica de barras
    plt.bar(campeones, frecuencias, color='skyblue')
    plt.xlabel('Campeones')
    plt.ylabel('Frecuencia de Uso')
    plt.title('Campeones Mas Usados')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    
    # Se crean cadenas por si se pide guardar la grafica
    carpeta_graficas = 'Graficas'
    prefijo_archivo = f"grafico-personaje-mas-jugado-{nombre}"
    patron_archivo = f"{prefijo_archivo}.*"  

    try:
        # Si el usuario elige guardar la funcion o sea(a o b)
        if menu_grafica == 'a' or menu_grafica == 'b':
            
            # Se crea la carpeta si no existe
            os.makedirs(carpeta_graficas, exist_ok=True)
            
            # Obtener el siguiente numero para el nombre del archivo
            siguiente_numero = obtener_siguiente_numero_archivos(carpeta_graficas, patron_archivo)
            
            # Obtener la fecha y hora actual para el nombre del archivo
            fecha_hora_actual = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            
            # Guardar la figura en la carpeta
            ruta_archivo = os.path.join(carpeta_graficas, f"{prefijo_archivo}({siguiente_numero})_{fecha_hora_actual}.png")
            plt.savefig(ruta_archivo)
            print(F"\033[92mGRAFICA GUARDADA EXITOSAMENTE EN {ruta_archivo}\033[0m")
            print("")
        if menu_grafica == 'a':
          plt.show()
        plt.close()

    except Exception as e:
        print(f"Error al guardar la grafica: {e}")

    if menu_grafica == 'a':
        plt.show()
        
# Funcion que guarda las consultas en un libro de excel
def guardar_consulta_api(url, carpeta_graficas='Consulta API'):
    try:
        # Primero obtiene una cadena de fecha y hora actual por separado
        fecha_actual = datetime.now().strftime("%d-%m-%Y")
        hora_actual = datetime.now().strftime("%H:%M:%S")
        
        # Despues se crea la ruta del archivo xlsx
        prefijo_archivo = "consulta"
        ruta_archivo_xlsx = os.path.join(carpeta_graficas, f"{prefijo_archivo}_{fecha_actual}.xlsx")

        # Crear un nuevo libro de Excel o cargar uno existente
        if os.path.exists(ruta_archivo_xlsx):
            workbook = openpyxl.load_workbook(ruta_archivo_xlsx)
        else:
            workbook = openpyxl.Workbook()

        # Eliminar la hoja predeterminada si existe
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        # aqui se verifica si la hoja "Consultas" ya existe
        sheet_name = "Consultas"
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)

        # se obtiene la hoja "Consultas"
        sheet = workbook[sheet_name]

        # se agrega un renglon para separar las consultas
        sheet.append([""])

        # Escribir la informacion en la hoja
        sheet.append(["Consulta API", f"Fecha: {fecha_actual}", f"Hora: {hora_actual}"])
        sheet.append([f"Consulta hecha a: {url}"])
        
        # Aqui se ajusta automaticamente el ancho de las columnas al contenido
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # por ultimo se Guarda el archivo Excel
        workbook.save(ruta_archivo_xlsx)

    except Exception as e:
        print(f"Error al hacer la consulta API: {e}")

# Funcion de lectura de archivos para mostrar reportes
def imprimir_reportes():
    carpeta_reportes = 'Reportes'
    try:
        if not os.path.exists(carpeta_reportes):
            print("No hay reportes disponibles.")
            input("Continuar?")
            return
        archivos_reportes = [f for f in os.listdir(carpeta_reportes) if os.path.isfile(os.path.join(carpeta_reportes, f))]

        if not archivos_reportes:
            print("No hay reportes disponibles.")
            input("Continuar?")
            return
        print("Reportes disponibles:")
        for i, archivo in enumerate(archivos_reportes, start=1):
            print(f"{i}. {archivo}")

        try:
            opcion = int(input("\033[96mSeleccione el numero del reporte que desea imprimir (0 para salir): \033[0m"))
        except ValueError:
            print("Entrada invalida. Introduzca un numero valido.")
            return

        if opcion == 0:
            print("Saliendo.")
            return
        elif 1 <= opcion <= len(archivos_reportes):
            ruta_archivo = os.path.join(carpeta_reportes, archivos_reportes[opcion - 1])

            # Imprimir el contenido del archivo seleccionado
            with open(ruta_archivo, 'r', encoding='utf-8') as archivo:
                contenido = archivo.read()
                print(contenido)
                input("Continuar?")
        else:
            print("Opcion no valida. Introduzca un numero valido.")
    except OSError as e:
        print(f"Error al acceder a la carpeta de reportes: {e}")
    except Exception as e:
        print(f"Error inesperado: {e}")

